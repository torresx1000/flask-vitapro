import os
import uuid
from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import func

from app.db.database import session
from app.models import Despacho


def preview_excel(file_obj, upload_folder, allowed_extensions, max_rows=30):
    """
    Lee el archivo Excel sin modificarlo y devuelve datos para vista previa:
    nombres de hojas y primeras filas de la primera hoja.
    Realizado en función aparte para no alterar el flujo de procesamiento.
    """
    filename = file_obj.filename or ""
    if (
        not filename
        or "." not in filename
        or filename.rsplit(".", 1)[1].lower() not in allowed_extensions
    ):
        raise ValueError("Solo se permiten archivos Excel (.xlsx y .xlsm)")

    os.makedirs(upload_folder, exist_ok=True)
    ext = filename.rsplit(".", 1)[1].lower()
    temp_name = f"preview_{uuid.uuid4().hex}.{ext}"
    path = os.path.join(upload_folder, temp_name)

    try:
        file_obj.save(path)

        if ext == "xlsm":
            wb = load_workbook(path, keep_vba=False, data_only=True)
        else:
            wb = load_workbook(path, data_only=True)

        hojas = wb.sheetnames
        if not hojas:
            wb.close()
            return {"hojas": [], "preview": [], "hoja_actual": None}

        sheet = wb[hojas[0]]
        preview = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            preview.append([_cell_to_preview(c) for c in row])

        wb.close()
        return {
            "hojas": hojas,
            "preview": preview,
            "hoja_actual": hojas[0],
        }
    finally:
        if os.path.exists(path):
            try:
                os.remove(path)
            except OSError:
                pass


def _cell_to_preview(value):
    """Convierte valor de celda a tipo serializable para JSON."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()


def procesar_excel(file_obj, form_data, upload_folder, allowed_extensions):

    filename = file_obj.filename or ""

    if (
        not filename
        or "." not in filename
        or filename.rsplit(".", 1)[1].lower() not in allowed_extensions
    ):
        raise ValueError("Solo se permiten archivos Excel (.xlsx y .xlsm)")

    os.makedirs(upload_folder, exist_ok=True)

    path = os.path.join(upload_folder, filename)

    file_obj.save(path)

    ext = filename.rsplit(".", 1)[1].lower()

    # decidir qué proceso usar
    if ext == "xlsm":
        return procesar_excel_xlsm(path, form_data)
    else:
        return procesar_excel_xlsx(path, form_data)


# -----------------------------
# PROCESAR XLSX
# -----------------------------
def procesar_excel_xlsx(path, form_data):

    wb = load_workbook(path)

    procesar_datos(wb, form_data)

    wb.save(path)
    wb.close()

    return path


# -----------------------------
# PROCESAR XLSM (CON MACROS)
# -----------------------------
def procesar_excel_xlsm(path, form_data):

    wb = load_workbook(path, keep_vba=True)

    procesar_datos(wb, form_data)

    wb.save(path)
    wb.close()

    return path


# -----------------------------
# LOGICA DE PROCESAMIENTO
# -----------------------------
def procesar_datos(wb, form_data):

    col_codigo = form_data["col_codigo"].upper()
    fila_inicio = int(form_data["fila_inicio"])
    fila_final = int(form_data["fila_final"])
    col_resultado = form_data["col_resultado"].upper()
    hoja_nombre = (form_data.get("hoja") or "").strip()
    if not hoja_nombre and wb.sheetnames:
        hoja_nombre = wb.sheetnames[0]

    if hoja_nombre not in wb.sheetnames:
        raise KeyError(f"La hoja no existe. Hojas disponibles: {wb.sheetnames}")

    sheet = wb[hoja_nombre]

    # 🔹 construir diccionario limpio desde la base de datos
    resultados = {
        str(codigo).strip(): float(valor) if valor is not None else 0
        for codigo, valor in session.query(
            Despacho.codigo,
            func.sum(Despacho.despacho * Despacho.peso_promedio),
        )
        .group_by(Despacho.codigo)
        .all()
    }

    # 🔹 recorrer filas del Excel
    for fila in range(fila_inicio, fila_final + 1):

        codigo = sheet[f"{col_codigo}{fila}"].value

        if not codigo:
            continue

        # normalizar código para evitar errores de tipo
        codigo = str(codigo).strip()

        if codigo in resultados:
            sheet[f"{col_resultado}{fila}"] = resultados[codigo]

def exportar_excel():

    registros = session.query(Despacho).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Registros"

    headers = [
        "ID",
        "ID Carga",
        "Fecha",
        "Turno",
        "Codigo",
        "Lote",
        "Peso Promedio",
        "Ubicacion",
        "Stock Inicial",
        "Cantidad SAP",
        "Despacho",
        "Saldo",
        "Observaciones",
    ]

    ws.append(headers)

    for r in registros:
        ws.append([
            r.id,
            r.id_carga,
            r.fecha,
            r.turno,
            r.codigo,
            r.lote,
            r.peso_promedio,
            r.ubicacion,
            r.stock_inicial,
            r.cantidad_sap,
            r.despacho,
            r.saldo,
            r.observaciones,
        ])

    bio = BytesIO()

    wb.save(bio)
    wb.close()

    bio.seek(0)

    return bio.getvalue()