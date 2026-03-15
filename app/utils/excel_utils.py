from openpyxl import load_workbook


def load_workbook_keep_vba(path: str, **kwargs):
    """Load an Excel workbook preserving macros if present."""
    return load_workbook(path, keep_vba=True, **kwargs)